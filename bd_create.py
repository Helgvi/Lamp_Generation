import openpyxl
import sqlite3


path_bd = 'db.sqlite'
con = sqlite3.connect(path_bd)
cur = con.cursor()
index = 0


def create_bd():
    cur.execute('''
      CREATE TABLE IF NOT EXISTS lamps(
        id INTEGER PRIMARY KEY,
        brand TEXT,
        model TEXT,
        eyars TEXT,
        lamp TEXT
        );
    ''')


def import_table_to_bd():
    cur = con.cursor()
    wb = openpyxl.load_workbook(filename="C:/h4-s.xlsx", read_only=True)
    list_name = wb.sheetnames
    sheet = wb[list_name[0]]
    i = index
    num = len(tuple(sheet.rows))
    for row in sheet.iter_rows(min_row=1,
                               max_row=num,
                               min_col=1,
                               max_col=4,
                               values_only=True):
        i = i + 1
        line = (i, row[0], row[1], row[2], row[3])
        print(line)
        cur.execute(
            'INSERT INTO lamps VALUES(?, ?, ?, ?, ?);',
            line
        )
    con.commit()
    con.close()


def main():
    create_bd()
    import_table_to_bd()


main()
